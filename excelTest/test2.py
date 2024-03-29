from openpyxl import Workbook, load_workbook

wb1 = Workbook()
wb = load_workbook('E:\python_workspace\Send-emails-automatically-in-batches\emailTest.xlsx')

ws = wb['资源研发中心【杭州办公区VPN网络申请单】']
for sheet in wb:
    print(sheet.title)

# openpyxl只能处理 .xlsx 合适的表格
# 获取所有工作表名
names = wb.sheetnames
# wb.get_sheet_by_name(name) 已经废弃,使用wb[name] 获取指定工作表
sheet = wb[names[0]]
# 获取最大行数
maxRow = sheet.max_row
# 获取最大列数
maxColumn = sheet.max_column
# 获取当前活动表
current_sheet = wb.active
# 获取当前活动表名称
current_name = sheet.title
# 通过名字访问Cell对象, 通过value属性获取值
a1 = sheet['A1'].value
# 通过行和列确定数据
a12 = sheet.cell(row=1, column=2).value
# 获取列字母
# column_name = utils.cell.get_column_letter(1)
# 将列字母转为数字, 参数忽略大小写
# column_name_num = openpyxl.utils.cell.column_index_from_string('a')
# 获取一列数据, sheet.iter_rows() 获取所有的行
"""
(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>)
(<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>)
(<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>)
(<Cell 'Sheet1'.A4>, <Cell 'Sheet1'.B4>, <Cell 'Sheet1'.C4>)
(<Cell 'Sheet1'.A5>, <Cell 'Sheet1'.B5>, <Cell 'Sheet1'.C5>)
"""
for one_column_data in sheet.iter_rows():
    print(one_column_data[0].value)

# 获取一行数据, sheet.iter_cols() 获取所有的列
"""
(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>)
(<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>)
(<Cell 'Sheet1'.C1>, <Cell 'Sheet1'.C2>, <Cell 'Sheet1'.C3>)
"""
for one_row_data in sheet.iter_cols():
    print(one_row_data[0].value, end="\t")

print("row = {}, column = {}".format(maxRow, maxColumn))
