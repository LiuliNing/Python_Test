# import openpyxl
#
# # xlsx = xlrd.open_workbook('C:/Users/Administrator/Desktop/test.xlsx')
# # print('All sheets: %s' % xlsx.sheet_names())
# wb = openpyxl.load_workbook('C:/Users/Administrator/Desktop/test.xlsx')
# sh = wb['边坡说明']
# ce = sh.cell(row=1, column=1)
# print(ce.value)
# print(list(sh.rows)[1:])
# for cases in list(sh.rows)[1:]:
#     case_id = cases[0].value
#     case_excepted = cases[1].value
#     case_data = cases[2].value
#     print(case_excepted, case_data)
# # 关闭工作薄
# wb.close()
import openpyxl


class Case:
    pass


class ReadExcel(object):
    def __init__(self, filename, sheetname):
        self.wb = openpyxl.load_workbook(filename)
        self.sh = self.wb[sheetname]

    def read_data_obj(self):
        """
        按行读取数据  每条用例存储在一个对象中
        :return:
        """
        rows_data = list(self.sh.rows)
        # print(rows_data)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        # print(titles)
        # 定义一个空列表用来存储测试用例
        cases = []
        for case in rows_data[1:]:
            # print(case)
            # 创建一个Case类的对象，用来保存用例数据
            case_obj = Case()
            data = []
            for cell in case:  # 获取一条测试用例数据
                # print(cell.value)
                # data.append(cell.value)
                # print(data)
                data.append(cell.value)
                # if isinstance(cell.value,str):  # 判断该单元格是否为字符串，如果是字符串类型则需要使用eval();如果不是字符串类型则不需要使用eval()
                #     data.append(cell.value)
                #     data.append(eval(cell.value))
                # else:
                #     data.append(cell.value)
            # 将该条数据存放至cases中
            # print(dict(list(zip(titles,data))))
            case_data = list(zip(titles, data))
            # print(case_data)
            for i in case_data:
                setattr(case_obj, i[0], i[1])
            # print(case_obj)
            # print(case_obj.case_id,case_obj.data,case_obj.excepted)
            cases.append(case_obj)
        return cases


if __name__ == '__main__':
    filePath = "C:/Users/Administrator/Desktop/理正接口格式信息整理-lln.xlsx"
    sheetName = "钻孔数据$ZK"
    r = ReadExcel(filePath, sheetName)
    res = r.read_data_obj()
    for i in res:
        print(i.caseid, i.excepted, i.data)
